import pandas as pd
from info import dist_list
import streamlit as st
from io import BytesIO
import numpy as np
import base64
from datetime import datetime
from openpyxl import load_workbook
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from github import Github


GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
GITHUB_REPO = "mustafasava/Sales_repo"

g = Github(GITHUB_TOKEN)
repo = g.get_repo(GITHUB_REPO)


def check_missing(prep_df,dist_name,year,month):
    try:
        mapping_file = f"./mapping/map_{dist_name}.xlsx"

        products = pd.read_excel(mapping_file, sheet_name="products")
        bricks = pd.read_excel(mapping_file, sheet_name="bricks", dtype={"dist_brickcode":str})

        merged_products = prep_df.merge(
            products,
            left_on="item_code",
            right_on="dist_itemcode",
            how="left"
        )

        missed_products = merged_products[merged_products["dist_itemcode"].isna()][["item_code","item_name","item"]].drop_duplicates()

        if not missed_products.empty:

            products_list = pd.read_excel("./mapping/main_lists.xlsx",sheet_name="products")[["product","barcode"]]
            name_to_code = dict(zip(products_list["product"], products_list["barcode"]))
            disabled_colsp = [col for col in missed_products.columns if col != "item"]

            st.write("### Enter missing Products mappings")
            st.data_editor(missed_products,column_config={"item": st.column_config.SelectboxColumn("item",options=list(name_to_code.keys()),
            required=True)}, disabled=disabled_colsp ,hide_index=True)
        else:
            st.success("No missing products")

        
        merged_bricks = prep_df.merge(
            bricks,
            left_on="brick_code",
            right_on="dist_brickcode",
            how="left"
        )

        missed_bricks = merged_bricks[merged_bricks["dist_brickcode"].isna()][dist_list[dist_name][2]+["brick"]].drop_duplicates()

        if not missed_bricks.empty:

            bricks_list = pd.read_excel("./mapping/main_lists.xlsx",sheet_name="bricks")
            

            disabled_colsb = [col for col in missed_bricks.columns if col != "brick"]
            st.write("### Enter missing Bricks mappings")

            missing_bricks = st.data_editor(missed_bricks,column_config={"brick": st.column_config.SelectboxColumn("brick",options=[""]+bricks_list["bricks"].tolist(),
                                required=True)},disabled=disabled_colsb,hide_index=True)


            if st.button("save_bricks"):
                missing_bricks = missing_bricks.drop_duplicates(subset=["brick_code"])
                missing_bricks = missing_bricks.dropna(subset=["brick"],how = "all")
                missing_bricks = missing_bricks.rename(columns={"brick_code":"dist_brickcode"})
                missing_bricks["added_by"] = st.session_state.get("username")
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                missing_bricks["date_time"] =  timestamp

                missing_bricks = missing_bricks[["dist_brickcode","brick","added_by","date_time"]]
                new_mapped_bricks = pd.concat([bricks, missing_bricks], ignore_index=True).drop_duplicates(subset=["dist_brickcode"])
                replace_sheet_in_github(new_mapped_bricks, "bricks", mapping_file, "Update bricks")


        else:
            st.success("No missing bricks")

        
        if missed_bricks.empty and missed_products.empty:
            final_merged = prep_df.merge(
                products,
                left_on="item_code",
                right_on="dist_itemcode",
                how="left"
            ).merge(
                bricks,
                left_on="brick_code",
                right_on="dist_brickcode",
                how="left"
            )
            return final_merged , dist_name, year ,month
        
        else:
            return 
    except Exception as e:
        st.error(f"{e}")   



def replace_sheet_in_github(df, sheet_name, path_in_repo, commit_msg):

    # --- Step 1: Get current file from GitHub ---
    file = repo.get_contents(path_in_repo)
    content = base64.b64decode(file.content)
    buffer = BytesIO(content)

    # --- Step 2: Load workbook and replace sheet ---
    wb = load_workbook(buffer)

    if sheet_name in wb.sheetnames:
        # remove old sheet
        std = wb[sheet_name]
        wb.remove(std)

    # create new sheet with updated df
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # --- Step 3: Save updated workbook to memory ---
    new_buffer = BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)
    new_content = base64.b64encode(new_buffer.read()).decode()

    # --- Step 4: Push updated file to GitHub ---
    repo.update_file(path_in_repo, commit_msg, new_content, file.sha, branch="main")

    st.success(f"Replaced sheet '{sheet_name}' in {path_in_repo}")